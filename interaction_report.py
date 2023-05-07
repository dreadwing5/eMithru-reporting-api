# report_generator.py

from typing import List, Dict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


class ExcelReportGenerator:
    def __init__(self, data: List[Dict[str, any]], filename: str):
        self.data = data
        self.filename = filename
        self.df = pd.DataFrame(data)
        self.wb = Workbook()
        self.ws = self.wb.active

    def reindex_and_rename_columns(self, columns_order: List[str], columns_mapping: Dict[str, str]):
        self.df = self.df.reindex(columns=columns_order)
        self.df.rename(columns=columns_mapping, inplace=True)
        return self

    def apply_datetime_conversion(self, datetime_columns: List[str], date_format: str = None):
        for column in datetime_columns:
            if column in self.df.columns:
                self.df[column] = self.df[column].apply(
                    lambda x: pd.to_datetime(
                        x, format=date_format, errors='ignore') if pd.notnull(x) else None
                )
        return self

    def _apply_date_styles(self):
        created_at_col = get_column_letter(
            self.df.columns.get_loc('Created At') + 1)
        closed_at_col = get_column_letter(
            self.df.columns.get_loc('Closed At') + 1)
        date_style = NamedStyle(name="date_style")
        date_style.number_format = 'MM/DD/YYYY'
        date_alignment = Alignment(horizontal='left')

        for cell in self.ws[created_at_col][1:]:
            if cell.value is not None:
                cell.style = date_style
                cell.alignment = date_alignment

        for cell in self.ws[closed_at_col][1:]:
            if cell.value is not None:
                cell.style = date_style
                cell.alignment = date_alignment

    def _apply_header_styles(self, header_style):
        for col in range(1, len(self.data[0]) + 1):
            cell = self.ws[get_column_letter(col) + '1']
            cell.style = header_style

    def _set_column_width_and_height(self):
        for row in self.ws.iter_rows(min_row=2):
            self.ws.row_dimensions[row[0].row].height = 75

        desc_col = get_column_letter(
            self.df.columns.get_loc('Description') + 1)
        for cell in self.ws[desc_col][1:]:
            cell.alignment = Alignment(
                horizontal='left', vertical='top', wrap_text=True, shrink_to_fit=True)

        for col in self.ws.columns:
            max_length = 0
            column = col[0].column_letter

            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except TypeError:
                    pass

            adjusted_width = max_length + 10
            self.ws.column_dimensions[column].width = min(adjusted_width, 60)

    def _create_table(self):
        table = Table(
            ref=f"A1:{get_column_letter(len(self.data[0]))}{len(self.df) + 1}", displayName="DataTable")
        table_style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                     showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = table_style
        self.ws.add_table(table)

    def create_excel_report(self):
        header_style = NamedStyle(name="header")
        header_style.font = Font(bold=True, color='ffffff')
        header_style.fill = PatternFill(fill_type='solid', fgColor='0077b6')

    # Convert the DataFrame to rows and write to the worksheet
        for r in dataframe_to_rows(self.df, index=False, header=True):
            self.ws.append(r)

        self._apply_date_styles()
        self._set_column_width_and_height()
        self._create_table()

    # Save the workbook to an Excel file
        self.wb.save(self.filename)
        return self
