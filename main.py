from typing import Union

from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import Dict, List
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from fastapi.middleware.cors import CORSMiddleware

app = FastAPI()


origins = [
    "http://localhost",
    "http://localhost:3000",
    "http://localhost:8000",
    "http://localhost:8080",
    "https://report-generator-api.onrender.com"
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)



class Data(BaseModel):
    data: Dict[str, List[int]]

def create_styled_excel(data: List[Dict[str, any]], filename: str) -> None:
    df = pd.DataFrame(data)
    df = df.reindex(columns=['title', 'tag', 'participants', 'state', 'openDate', 'closeDate', 'description'])

    header_style = NamedStyle(name="header")
    header_style.font = Font(bold=True, size=16)
    header_style.alignment = Alignment(horizontal='center')
    header_style.fill = PatternFill(fill_type='solid', fgColor='FFFF00')


    # Create a new workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Convert the DataFrame to rows and write to the worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Apply the header style to the header row
    for col in range(1, len(data[0]) + 1):
        cell = ws[get_column_letter(col) + '1']
        cell.value = cell.value.upper()
        cell.style = header_style

    # Set the width and height of each column based on the content
    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = 75

    desc_col = get_column_letter(df.columns.get_loc('description') + 1)
    for cell in ws[desc_col][1:]:
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, shrink_to_fit=True)

    for col in ws.columns:
            max_length = 0
            column = col[0].column_letter

            # Find the maximum length of the content in the column
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except TypeError:
                    pass

            # Set the width of the column based on the maximum length of the content
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = min(adjusted_width, 100)

    # Save the workbook to an Excel file
    wb.save('data.xlsx')


@app.post("/generate_excel")
async def generate_excel(data: List[Dict]):
    try:
        create_styled_excel(data, "data.xlsx")
        return FileResponse("data.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename="data.xlsx")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8080)
