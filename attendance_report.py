from typing import Dict, List
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font


class AttendanceReportGenerator:
    def __init__(self, data: List[Dict], filename: str):
        attendance_data = data["attendanceData"]["subjects"]
        self.df = pd.DataFrame(attendance_data)
        self.filename = filename
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Attendance Report"

    def generate_pivot_table(self):
        pivot_table = pd.pivot_table(self.df, values=["attendedClasses", "totalClasses"],
                                     index=["subjectCode", "subjectName"],
                                     aggfunc={"attendedClasses": "sum", "totalClasses": "sum"})
        pivot_table.reset_index(inplace=True)
        pivot_table.columns = [
            "Subject Code", "Subject Name", "Attended Classes", "Total Classes"]
        pivot_table["Percentage"] = (
            pivot_table["Attended Classes"] / pivot_table["Total Classes"]) * 100

        self.pivot_table = pivot_table

    def calculate_overall_attendance(self):
        overall_attendance = (self.pivot_table["Attended Classes"].sum() /
                              self.pivot_table["Total Classes"].sum()) * 100
        self.overall_attendance = overall_attendance

    def write_to_worksheet(self):
        for r in range(len(self.pivot_table.index) + 1):
            for c in range(len(self.pivot_table.columns)):
                self.ws.cell(row=r + 1, column=c + 1,
                             value=self.pivot_table.iat[r - 1, c] if r > 0 else self.pivot_table.columns[c])

        header_font = Font(bold=True)
        for cell in self.ws[1]:
            cell.font = header_font

        self.ws.cell(row=len(self.pivot_table.index) + 3,
                     column=1, value="Overall Attendance")
        self.ws.cell(row=len(self.pivot_table.index) + 3, column=2,
                     value=f"{self.overall_attendance:.2f}%")

    def save_report(self):
        self.wb.save(self.filename)

    def generate_report(self):
        self.generate_pivot_table()
        self.calculate_overall_attendance()
        self.write_to_worksheet()
        self.save_report()
